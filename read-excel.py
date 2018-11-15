
# you need to install the following modules using pip
# you can 
# commands to install  openpyxl, numpy and pandas
#sudo pip install openpyxl
#sudo pip install numpy
#sudo pip install pandas

#this program with simply  print it on the screen
# make sure you save the excel file in the latest xlsx format not the old  xls format
# if you run it as "python read-excel.py" 
# so you can redirect to a file such as  "python read-excel.py > file.tsv"
# you can open it using any excel or the libreoffice in ubuntu and 
# then save it as excel once you open it and you can format the date column as you want 


# import this module 
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill

#name of the excel sheet  change to where you have your file or
# make it a command line option
wb = load_workbook(filename = '/home/ubuntu/first_input_file.xlsx')

#sheet I am going to read from 
sheet_ranges = wb['Sheet1']

#wb1 = open_workbook(filename = '/home/ubuntu/output_file.xlsx')
wb1 = Workbook()
ws1 = wb1.create_sheet(title="MySheet")

validCell =True
i=2


# we will store in the following format
#date["24/02/17"] = [3000, 234.32, 123.89]
#date["05/05/17"] = [23333, 4343, 43.55]
#.......
#.......
data = {}

# read the data into data dictionary
while True:  #loop for ever
  checkvalue = sheet_ranges['A'+ str(i)].value
  if checkvalue==None:   #no data , make sure there is no invisibe tab or spaces, if empty quit
     break

  no = str(sheet_ranges['A'+ str(i)].value).strip()  # make a string

  _date = sheet_ranges['B'+ str(i)].value 
  _amount= sheet_ranges['C'+ str(i)].value
  date = str(_date).strip()  # make sure you remove invisible spaces like tag, spaces 
  try:
      amount = float(str(_amount).strip() ) # first make sure it is a string, make sure you remove invisible spaces like tag, spaces 
                                 # and then convert to number
  except:
      amount =0    # if for some reason you cannot convert to a number, assmume it to be a zero 
      print "error amunt  in row ", i                 # issue some error messages on the screen

  if not date in data:  # have not seen this date before
     data[date] = []   #  initialize an empty array for this

  # now safely without worrying insert the amount
  data[date].append(amount)
  i += 1

# you can uncomment and check to print
#print data

i=1
#one date at a time
for onedate in data:
    # for each data keep summing and printing one the screen
    sum = 0
    # one deposit for that date
    for amount in data[onedate]:
        sum += amount 
        print str(i) + "\t" + onedate + "\t" +  str(amount)
        ws1["A"+ str(i)] = str(i)   # + "\t" + onedate + "\t" +  str(amount)
        ws1["B"+ str(i)] = onedate 
        ws1["C"+ str(i)] =  str(amount)
        i +=1  #increment row no

    #now print the sum of the date
    print "" + "\t" + "Total" + "\t" +  str(sum)
    ws1["A"+ str(i)] = ""   # + "\t" + onedate + "\t" +  str(amount)
    ws1["B"+ str(i)] = "Subtotal"
    ws1["C"+ str(i)] =  str(sum)

    ws1['B' + str(i)].fill = PatternFill(fgColor="FFEE08", fill_type = "solid")
    ws1['C' + str(i)].fill = PatternFill(fgColor="823030", fill_type = "solid")

    i += 1

    
wb1.save(filename = '/home/ubuntu/output_file.xlsx')
